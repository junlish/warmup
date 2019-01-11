#coding=utf-8
'''
  检查PC的领用、借用、归还、入库等事件记录
@author: junli
'''

import datetime
import os

from collections import defaultdict
from operator import  attrgetter
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Border,PatternFill,Side
from openpyxl.utils import get_column_letter




class EventRecordRow:
    
    headers = "日期    变更类型    单据编号    品牌    型号    序列号    快速服务代码    使用部门    使用人    备注".split()
    
    @classmethod
    def header_length(cls):
        return len(cls.headers)
    
    def __init__(self,values):
        self.values = values
        
    @property
    def event_date(self):
        return str(self.values[0])

    @property
    def seq_no(self):
        return self.values[5]
    
    @property
    def sheet_no(self):
        return self.values[2]
    
    @property
    def event_type(self):
        return self.values[1]


class EventsValidator:
    NEXT_EVENTS = {
            '借用' : ['归还', '入库'],
            '领用' : [ '入库'],
            '归还': ['借用', '领用'],
            '入库': ['借用', '领用']
        }
    
    @staticmethod
    def is_events_valide(event_types):
        if len(event_types) > 1:
            for i in range(len(event_types)-1):
                next_event = event_types[i+1]
                if next_event not in EventsValidator.NEXT_EVENTS[event_types[i]]:
                    return False
        return True


class PCEventsStore:
    
    def __init__(self):
        self.seq_events = {}
        self.sheet_events = {}
        self.incomplete_events = []
        self.count = 0
        
    
    def load_from_file(self,filepath):
        wb = load_workbook(filepath)
        ws = wb['变更记录']
        
        sheet_events = defaultdict(list)
        seq_events = defaultdict(list)
        incomplete_events = []
        record_col_count = EventRecordRow.header_length()
        count = 0
        for row in ws.rows:           
            # skip header row
            if row[0].value=='日期':
                continue                 
            # empty row: stop
            if row[0].value is None and row[1].value is None and row[2].value is None:
                break            
            
            values = [cell.value for cell in row[:record_col_count]]
            record =  EventRecordRow(values)
            # if no seq_no, bad rows 
            if record.seq_no is None or record.seq_no=='未带待核实':
                incomplete_events.append(record)
            else:
                seq_events[record.seq_no].append(record)
            
            if record.sheet_no is not None:
                sheet_events[record.sheet_no].append(record)            
            count +=1                    
                
        self.sheet_events = sheet_events
        self.seq_events = seq_events
        self.incomplete_events = incomplete_events
        self.count = count


    
    def validate_events_sequences(self):       
        bad_pattern_seqs = defaultdict(list)
        lack_sheet_seqs = []
        for seq,events in self.seq_events.items():
            # sort by date
            events.sort(key=attrgetter('event_date'))
            event_types = [event.event_type for event in events]
            if not EventsValidator.is_events_valide(event_types):
                pattern = "-".join(event_types)
                bad_pattern_seqs[pattern].append(seq)
            
            last_event = events[-1]
            if event_types[-1] in ['借用', '领用'] and last_event.sheet_no is None:
                lack_sheet_seqs.append(last_event)
                
                
        return bad_pattern_seqs,lack_sheet_seqs
    
    def machines_in_store(self):
        seqs_in_store = []
        for _,events in self.seq_events.items():
            # sort by date
            events.sort(key=attrgetter('event_date'))
            event_types = [event.event_type for event in events]
            
            if event_types[-1] in ['归还', '入库']:
                seqs_in_store.append(events[-1])
        return seqs_in_store
                
       
def main():
    filepath = r'D:\shadow\办公支持\办公电脑\办公电脑实物台帐.xlsx'
    base_name ="pcs-"+ datetime.datetime.now().strftime('%Y%m%d_%H%M%S') +".xlsx"
    outputpath = os.path.join(r'c:\temp', base_name)
    store = PCEventsStore()
    store.load_from_file(filepath)
    print("Load {} events from file {}; Bad records: {}".format(store.count, filepath, len(store.incomplete_events)))
    print("... about {} machines.".format(len(store.seq_events)))
    
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Load {} events from file {}".format(store.count, filepath)])
    ws_summary.append(["... about {} machines.".format(len(store.seq_events))])
    ws_summary.append(["","", "Generated @ "+ datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')])    
    ws_summary.append([])
    
    if len(store.incomplete_events)>0:
        ws_summary.append(["Bad Records:"])
        ws_summary.append(EventRecordRow.headers)   
        for e in store.incomplete_events:        
            ws_summary.append(e.values)
        
  
       
    ws_summary.append([])
    bad_pattern_seqs,lack_sheet_seqs = store.validate_events_sequences()
    ws_summary.append(["Bad sequence of events:"])
      
    
    print("Bad sequence of events: {}".format(len(bad_pattern_seqs)))
    ws_summary.append(EventRecordRow.headers)
    for p,seqs in bad_pattern_seqs.items():
        print("{} : {}".format(p, ','.join(seqs)))               
        for seq in seqs:
            events = store.seq_events[seq]
            for e in events:
                ws_summary.append(e.values)
    
    ws_no_sheetno = wb.create_sheet("NoSheetNo")
    ws_no_sheetno.append(EventRecordRow.headers)
    print("Those machines are lacking a sheet no : {}".format(len(lack_sheet_seqs)))
    for event in lack_sheet_seqs:
        ws_no_sheetno.append(event.values)

    ws_instore = wb.create_sheet("InStore")
    ws_instore.append(EventRecordRow.headers)
    events = store.machines_in_store()
    print("Machines in store: {}".format(len(events)))
    for s in events:
        ws_instore.append(s.values)
    
    wb.save(outputpath)
    print("Saved file to {}".format(outputpath))

if __name__ == '__main__':    
    main()