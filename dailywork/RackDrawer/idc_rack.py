#coding=utf-8
'''
Created on 2018年7月11日
    从数据excel表生成机柜落位图， 每个机柜列生成一个Excel Sheet.
    源文件表格格式： 设备类型    机柜位    机柜内位置    设备品牌    设备型号    设备序列号    固定资产条码    信息变更时间    备注
@author: junli
'''
import re
from collections import defaultdict
import csv
import datetime
import os.path
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font,Alignment,Border,PatternFill,Side
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook





class IDCRecordRow:
    headers = "设备类型    机柜位    机柜内位置    设备品牌    设备型号    设备序列号    固定资产条码    信息变更时间    备注".split()
    header_widths = [16,12,9,9,9,16,16,12,16]
    def __init__(self, values):
        self.values = list(values)
    
    @property    
    def rack(self):
        return self.values[1]
    
    def is_summary_item(self):
        return self.values[0].strip() == "存储设备" and (self.values[-1] is not None  and self.values[-1].strip()=="汇总记录")
      
    # supported patters:  12U |  12 | 12-13 | 13-12 | 12U-13U
    def get_u_range(self):
        u_no_range = self.values[2]
        u_no_range = u_no_range.upper().replace("U","")
        # Example: 12-13
        if u_no_range.find('-')>0:
            start, end = u_no_range.split('-')
            start = int(start.strip())
            end = int(end.strip())
            if start > end:
                start,end = end, start
        else:
            #  Example: 12
            start = int(u_no_range.strip())
            end = start
#        if end > MAX_U_COUNT:
#            raise ValueError("Invalid u_no：{} @ {}".format( self.values[2], self.rack))
        
        return start,end

    @classmethod
    def header_length(cls):
        return len(cls.headers)

'''
    Read Rack configuration from a csv file, expected records:
        RackName,UCount,1UPos,Comments
'''
class RacksConfig:
    
    def __init__(self, default_u_count=42, default_1U_from_bottom=True):
        self.racks = {}
        self.default_u_count = default_u_count
        self.default_1U_from_bottom = default_1U_from_bottom
        
    
    def load_from_file(self, filepath):
        with open(filepath) as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                self.racks[row['RackName']] = (int(row['UCount']), row['1UPos'].lower()=="bottom")
    
    '''
        Returns racks max u count and if 1u start from rack's bottom. 
        If no configuration exists for this rack, return a default configuration.
    '''                
    def _get_rack_config(self, rackname):
        if rackname in self.racks:
            return self.racks[rackname]
        else:
            return (self.default_u_count, self.default_1U_from_bottom)
    
    
    def get_rack_max_u_count(self,rackname):
        return self._get_rack_config(rackname)[0]
    
    # create a mapping for u# and row#;
    # start_row is the row# where put 1U
    def get_rack_row_mapping(self, rackname, start_row):
        rack_u_count, start_from_bottom = self._get_rack_config(rackname)
        u_to_row_mapping = {}
        for i in range(1,rack_u_count+1):
            if start_from_bottom:
                u_to_row_mapping[i] = rack_u_count - i + start_row
            else:
                u_to_row_mapping[i] = start_row+i-1
        
        return u_to_row_mapping
        


class SourceReader:
    
    def __init__(self):
        # rack name --> [item1, item2...]
        self.records_per_rack = {}
        self._item_count=0
        # column name --> [rack1, rack2...]
        self.col_rack_map = {}
        
    
    #read records from a xlsx file, and return a dict of rack --> list of items
    def load_from_file(self,filepath, sheet_name):
        wb = load_workbook(filepath, read_only=True)
        if sheet_name not in wb.sheetnames:
            sheet_name = 'Sheet1'   # using default sheet 'Sheet1'
        ws = wb[sheet_name]
        records_per_rack = defaultdict(list)
        count =0
        record_col_count = len(IDCRecordRow.headers)
        for row in ws.rows:
            # skip header row
            if row[0].value=='设备类型':
                continue     
                   
            record = IDCRecordRow(cell.value for cell in row[:record_col_count])
            if not record.is_summary_item():
                records_per_rack[record.rack].append(record)
                count +=1
        print("Load {} items in {} racks from file '{}'".format(count,len(records_per_rack), filepath ))
        
        # check overlap
      
        for rack_name in records_per_rack:
            u_range_list = []
            for record in records_per_rack[rack_name]:
                start,end = record.get_u_range()
                for r_start, r_end in u_range_list:
                    if (start >= r_start and start<=r_end)  or (end >=r_start and end <=r_end):
                        print("WARN: Rack {} has overlap:  [{}-{}] x [{}-{}] ".format(rack_name, start, end, r_start,r_end))
                u_range_list.append([start,end])
        
        self.records_per_rack = records_per_rack
        self._item_count = count

        racks = records_per_rack.keys()
        r = re.compile(r'(\w+-\w)(\d+)')
        col_rack_map = defaultdict(list)
        for rack_name in racks:
            match = r.match(rack_name)
            room_col,_ = match.groups()
            col_rack_map[room_col].append(rack_name)
        self.col_rack_map = col_rack_map
    
    @property
    def columns(self):
        return list(sorted(self.col_rack_map.keys()))
    
    def racks_in_col(self,col_name):
        return list(sorted(self.col_rack_map[col_name]))
    
    def items_in_rack(self, rack_name):
        return self.records_per_rack[rack_name]
    
    @property
    def item_count(self):
        return self._item_count
    
    @property
    def rack_count(self):
        return len(self.records_per_rack)
    
    def rack_used_u_count(self, rack_name):
        items = self.items_in_rack(rack_name)
        used_u_count = 0
        for item in items:
            start,end = item.get_u_range()
            used_u_count += end-start+1    
        return used_u_count
    
rack_sheet_pos = {}     
    
def render_column(ws, col_name,  reader, rackconfig ):
    print("Rendering Column {} ...".format(col_name))  
    racks = reader.racks_in_col(col_name)
    ws.column_dimensions[get_column_letter(1)].width = 3
    ws.column_dimensions[get_column_letter(1+ IDCRecordRow.header_length()+ 1)].width = 3
    for i,width in enumerate(IDCRecordRow.header_widths):
        ws.column_dimensions[get_column_letter(2+i)].width = width
    
    # draw each rack
    start_row = 1
    for i,rack_name in enumerate(list(sorted(racks))):
        rack_u_count = rackconfig.get_rack_max_u_count(rack_name)
        render_rack(ws, rack_name, reader,rackconfig, start_row, 1  )
        start_row = start_row + rack_u_count + 6
        


thin_border = Border(left=Side(style='thin'),  right=Side(style='thin'), 
                     top=Side(style='thin'),  bottom=Side(style='thin'))

def render_rack(ws, rack_name,  reader,  rackconfig,  ws_start_row=1, ws_start_col=1):
    records = reader.items_in_rack(rack_name)
    rack_render_start_pos = "{}{}".format(get_column_letter(ws_start_col),ws_start_row)
    rack_u_count = rackconfig.get_rack_max_u_count(rack_name)
    rack_sheet_pos[rack_name] = rack_render_start_pos
    print("...Rendering rack {} @ {} [{} items / {} U]".format(rack_name, rack_render_start_pos, len(records), 
                                                                     reader.rack_used_u_count(rack_name)))
    # Plan:
    # row 0 : merge[0 -- len(headers)+1] , rack_name
    # row 1: col headers
    # row 2 - MAX_U_COUNT+1:  MAX_U_COUNT U
    # col 0: U no#, 
    # col 1 -  len(headers): contents
    # col len(headers)+1: u no#
    #    if item span more than 1u, merge cells
    alcenter = Alignment(horizontal="center", vertical="center",wrap_text=True)
    header_fill = PatternFill("solid", fgColor='FFEFDB')
    
    # first draw header and u no columns
    header_len = IDCRecordRow.header_length()
    ws.merge_cells(start_row=ws_start_row, start_column=ws_start_col, end_row=ws_start_row, end_column= ws_start_col+ header_len+ 1  )
    header_cell = ws.cell(row=ws_start_row,column=ws_start_col, value= rack_name)
    header_font = Font(name='Bold', size=25)
    header_cell.font = header_font
    header_cell.alignment  = alcenter
    header_cell.fill = header_fill
    ws.row_dimensions[ws_start_row].height = 30
    
    # draw headers
    hrow = ws_start_row + 1
    for i in range(header_len):
        hcell = ws.cell(row=hrow,column=ws_start_col+1+i,value = IDCRecordRow.headers[i])
        tab_header_font= Font(name='Bold', size=11)
        hcell.font = tab_header_font
        hcell.alignment  = alcenter
        hcell.fill = header_fill
    ws.cell(row=ws_start_row+1, column=ws_start_col).fill = header_fill
    ws.cell(row=ws_start_row+1, column=ws_start_col+ header_len+ 1 ).fill = header_fill
        
    
   
    u_to_row_dict = rackconfig.get_rack_row_mapping(rack_name, ws_start_row+2)  
    
    u_no_fill = PatternFill("solid", fgColor='5CACEE')
    # draw u no#
    for i in range(1, rack_u_count+1):
        h_cell = ws.cell(row=u_to_row_dict[i],column=ws_start_col, value= i)
        t_cell = ws.cell(row=u_to_row_dict[i],column=ws_start_col+ header_len + 1, value= i)
        h_cell.fill = u_no_fill
        t_cell.fill = u_no_fill
    
      
       
    # draw items
    for item in records:
        start, end = item.get_u_range()        
        # safe guard: if end > max rack u count, print out an error and do not draw this item
        if end > rack_u_count:
            print(("Invalid u_no：{} @ {}".format( end, rack_name)))
            continue
            
        for tab_i in range(header_len):
            row_start, row_end = u_to_row_dict[start], u_to_row_dict[end]
            if row_start > row_end:
                row_start, row_end = row_end, row_start
            if start != end:
                # need to merge cells                    
                ws.merge_cells(start_row=row_start, start_column= ws_start_col+1+tab_i, end_row=row_end,end_column= ws_start_col+1+tab_i )
            value = item.values[tab_i]
            cell = ws.cell(row=row_start,column= ws_start_col+1+tab_i,value=value  )    
            if value is not None and isinstance(value, datetime.datetime):
                cell.number_format = 'yyyy-mm-dd'
            cell.alignment  = alcenter
            
    # set borders
    cell_range = ws[ "{}{}".format(get_column_letter(ws_start_col),ws_start_row): \
                     "{}{}".format(get_column_letter(ws_start_col+ header_len + 1),ws_start_row+rack_u_count+1)]
    
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for r in cell_range:
        for c in r:
            c.border = border 
            
    

def render(source_filepath, rackconfig_filepath, result_filepath, sheet_name):
    reader = SourceReader()
    reader.load_from_file(source_filepath, sheet_name)
    rackconfig = RacksConfig()
    rackconfig.load_from_file(rackconfig_filepath)
    columns = reader.columns
    print("Rack columns are:"+ str(columns))
       
    wb= Workbook()         
    for rack_col in columns:
        ws = wb.create_sheet(rack_col)
        #render_column(ws, rack_col, reader.racks_in_col(rack_col), records_per_rack )
        render_column(ws, rack_col, reader, rackconfig)
        
    ws = wb.active
    ws.title = "Summary"
    ws.append(["源文件:", source_filepath,"" ])
    ws.append(["","机柜数量：", reader.rack_count ])
    ws.append(["","设备数量：", reader.item_count ])
    ws.append([])
    
    ws.append(["","列",'机柜','设备数量', '占用U数'])
    start_row_no = ws.max_row
    for rack_col in columns:
        for rack_name in reader.racks_in_col(rack_col):
            used_u_no = reader.rack_used_u_count(rack_name)
            ws.append(["",rack_col, rack_name, len(reader.items_in_rack(rack_name)), used_u_no])
            rack_cell = ws.cell(ws.max_row, 3)
            rack_cell.hyperlink = "#'{}'!{}".format(rack_col, rack_sheet_pos[rack_name]) 
            rack_cell.style = 'Hyperlink'
            if used_u_no < 10:
                u_no_cell = ws.cell(ws.max_row, 5)
                u_no_cell.style = 'Warning Text'
    end_row_no = ws.max_row
    
    #set table borders
    for row_no in range(start_row_no, end_row_no+1):
        for col_no in range(2,6):
            c = ws.cell(row_no, col_no)
            c.border = thin_border
        
       
    ws.cell(ws.max_row+3, 1, "Generated @ "+ datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))    
    
    
    
    wb.save(result_filepath)
    print("Result saved to file {}.".format(result_filepath))


def main():
    parser = argparse.ArgumentParser(description='机柜落位图生成器')
    parser.add_argument('--source', required=True, help='源数据文件，期待数据表格式：源文件表格格式： 设备类型    机柜位    机柜内位置    设备品牌    设备型号    设备序列号    固定资产条码    信息变更时间    备注')
    parser.add_argument('--rackconfig', help='机柜最大U数，第1U的起始位置')
    parser.add_argument('--output',  help='输出文件; 默认为源文件同目录下的gen-yyyymmdd-HHMMSS.xlsx')   
    parser.add_argument('--sheet',  default='设备信息', help='输出文件; 默认为源文件同目录下的gen-yyyymmdd-HHMMSS.xlsx')   
    args = parser.parse_args()
    
    if not os.path.isfile(args.source):
        print("源文件{}不存在或不是文件.".format(args.source))
        exit(-1)
        
    if args.rackconfig is None:
        default_config_path = os.path.join(os.path.dirname(args.source), 'racks.csv')
        if not os.path.isfile(default_config_path):
            default_config_path =  os.path.join(os.path.dirname(os.path.realpath(__file__)), 'config', 'racks.csv')
        
        if not os.path.isfile(default_config_path):
            print("缺少机柜配置文件,也无法在默认文件夹找到配置文件")
            exit(-2)        
        
        args.rackconfig = default_config_path
        print("使用机柜配置文件：{}".format(default_config_path))
        
            
    
    if args.output is None:
        base_name ="gen-"+ datetime.datetime.now().strftime('%Y%m%d_%H%M%S') +".xlsx"
        args.output = os.path.join(os.path.dirname(args.source), base_name)
    
    render(args.source ,args.rackconfig, args.output, args.sheet, )
        
if __name__ == '__main__':
    main()
    # rackconfig = RacksConfig()
    #rackconfig.load_from_file("config\\racks.csv")
    #mapping =  rackconfig.get_rack_row_mapping("MD02-C02", 3)
    #print(mapping)
   
    
    
    